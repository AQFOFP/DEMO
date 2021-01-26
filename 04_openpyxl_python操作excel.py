

# openpyxl操作
import random

from mongoengine import connect, Document, StringField, IntField
from openpyxl import Workbook
from openpyxl.styles import Alignment

connect('mydb', host='localhost', port=27017)


class Ratain(Document):
    room_type = StringField(default='')
    second = IntField(default=0)
    third = IntField(default=0)
    date_str = StringField(default='')
    meta = {"strict": False}


def write_data():
    room_list = ['topic', 'music', 'game']
    date_str = ['2020-11-01', '2020-11-02', '2020-11-03']
    for r_type in room_list:
        for d_str in date_str:
            Ratain(room_type=r_type, date_str=d_str, second=random.randint(1, 10), third=random.randint(1, 10)).save()


def savexl():

    wb = Workbook()
    _save_sheet = wb.active

    rowNum = 1
    _save_sheet.cell(row=rowNum, column=1).value = 'name'
    _save_sheet.cell(row=rowNum, column=2).value = 'date_str'
    _save_sheet.cell(row=rowNum, column=3).value = 'second'
    _save_sheet.cell(row=rowNum, column=4).value = 'third'
    room_list = ['topic', 'music', 'game']
    rownum = 2
    for r_type in room_list:
        users = Ratain.objects(room_type=r_type).order_by('date_str').all()  # 返回所有的文档对象列表
        for index, item in enumerate(users):
            if index == 0:
                _save_sheet.cell(row=rownum, column=1).value = item.room_type
                _save_sheet.merge_cells(f"A{rownum}:A{rownum+2}")
                _save_sheet.cell(row=rownum, column=1).alignment = Alignment(horizontal="center", vertical="center")

            _save_sheet.cell(row=rownum, column=2).value = item.date_str
            _save_sheet.cell(row=rownum, column=3).value = item.second
            _save_sheet.cell(row=rownum, column=4).value = item.third
            rownum += 1

            print(item.room_type, item.second, item.third, item.date_str)

    path = 'list_credit_charge_bill.xlsx'
    # 把文件保存到linux本地，
    wb.save(path)


    # for index, item in enumerate(users):
    #     e_pos = index + 1
    #     _save_sheet.merge_cells("A1:A2")
    #     # _save_sheet.cell(row=rowNum + e_pos, column=1).value = item.name
    #     # _save_sheet.cell(row=rowNum + e_pos, column=2).value = item.age
    #
    # path = 'list_credit_charge_bill.xlsx'
    # # 把文件保存到linux本地，
    # wb.save(path)


if __name__ == '__main__':
    savexl()
    # write_data()